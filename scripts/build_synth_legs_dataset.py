#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import random
import shutil
import sys
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
TMP_DIR = ROOT / "synth_legs"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


COMMON_PROMPT = (
    "Count the number of legs in this animal glyph image. "
    "Answer with only a number in curly brackets, e.g., {4}. Do not add any other text."
)


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=(
            "Build synthetic bird+animal legs dataset (8192 default, 50/50 split), "
            "save parquet for training, and export 20-sample preview workbook."
        )
    )
    p.add_argument("--n-total", type=int, default=8192)
    p.add_argument("--seed", type=int, default=20260322)
    p.add_argument("--image-size", type=int, default=640)
    p.add_argument("--aa", type=int, default=3)
    p.add_argument(
        "--out-image-dir",
        type=Path,
        default=Path("data/generated/synth_legs/images"),
    )
    p.add_argument(
        "--out-parquet",
        type=Path,
        default=Path("data/generated/synth_legs/train_8192.parquet"),
    )
    p.add_argument(
        "--out-summary",
        type=Path,
        default=Path("data/generated/synth_legs/train_8192_summary.json"),
    )
    p.add_argument(
        "--sample-xlsx",
        type=Path,
        default=Path("data/generated/synth_legs/synth_legs_20_samples.xlsx"),
    )
    p.add_argument("--sample-n", type=int, default=20)
    p.add_argument("--no-shuffle", action="store_true")
    return p.parse_args()


def _import_generators():
    if str(TMP_DIR) not in sys.path:
        sys.path.insert(0, str(TMP_DIR))

    from synth_legs import generate_animals_synth_samples as animal_mod
    from synth_legs import generate_bird_synth_samples as bird_mod

    return bird_mod, animal_mod


def _clean_dir(path: Path) -> None:
    if path.exists():
        for p in path.glob("*"):
            if p.is_file():
                p.unlink()
            elif p.is_dir():
                shutil.rmtree(p)
    path.mkdir(parents=True, exist_ok=True)


def _crop_remove_label(src: Path, dst: Path, image_size: int) -> None:
    with Image.open(src).convert("RGB") as im:
        # Source from generator is (image_size, image_size + strip_h).
        cropped = im.crop((0, 0, image_size, image_size))
        dst.parent.mkdir(parents=True, exist_ok=True)
        cropped.save(dst, format="PNG")


def _build_rows(
    n_total: int,
    image_size: int,
    aa: int,
    out_image_dir: Path,
    seed: int,
) -> List[Dict[str, object]]:
    random.seed(seed)
    np.random.seed(seed)

    bird_mod, animal_mod = _import_generators()

    # Intermediate temp generation dirs (label-strip images), cleaned afterward.
    tmp_gen_root = ROOT / "data" / "generated" / "synth_legs_generation_tmp"
    tmp_bird = tmp_gen_root / "bird"
    tmp_animal = tmp_gen_root / "animal"
    _clean_dir(tmp_bird)
    _clean_dir(tmp_animal)

    bird_mod.OUT_DIR = tmp_bird
    animal_mod.OUT_DIR = tmp_animal

    n_bird = n_total // 2
    n_animal = n_total - n_bird

    rows: List[Dict[str, object]] = []

    # Bird rows.
    for i in range(n_bird):
        meta = bird_mod.make_one(i, size=image_size, aa=aa)
        src = tmp_bird / str(meta["file"])
        dst = out_image_dir / "bird" / f"{i:05d}_{meta['style']}_legs{meta['legs']}.png"
        _crop_remove_label(src, dst, image_size=image_size)
        src.unlink(missing_ok=True)
        rel_for_dataset = str(dst.relative_to(ROOT))

        rows.append(
            {
                "images": np.array([rel_for_dataset], dtype=object),
                "problem": f"<image>{COMMON_PROMPT}",
                "prompt": COMMON_PROMPT,
                "answer": str(int(meta["legs"])),
                "subset": "bird",
                "species": "bird",
                "style": str(meta["style"]),
                "legs": int(meta["legs"]),
                "image_rel_path": rel_for_dataset,
            }
        )

    # Animal rows (quadruped family).
    for i in range(n_animal):
        meta = animal_mod.make_one(i, size=image_size, aa=aa)
        src = tmp_animal / str(meta["file"])
        dst = out_image_dir / "animal" / f"{i:05d}_{meta['style']}_legs{meta['legs']}.png"
        _crop_remove_label(src, dst, image_size=image_size)
        src.unlink(missing_ok=True)
        rel_for_dataset = str(dst.relative_to(ROOT))

        rows.append(
            {
                "images": np.array([rel_for_dataset], dtype=object),
                "problem": f"<image>{COMMON_PROMPT}",
                "prompt": COMMON_PROMPT,
                "answer": str(int(meta["legs"])),
                "subset": "animal",
                "species": "quadruped",
                "style": str(meta["style"]),
                "legs": int(meta["legs"]),
                "image_rel_path": rel_for_dataset,
            }
        )

    shutil.rmtree(tmp_gen_root, ignore_errors=True)
    return rows


def _build_sample_excel(df: pd.DataFrame, xlsx_path: Path, sample_n: int) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    ws_cols = ["image_preview", "subset", "species", "style", "legs", "prompt", "answer", "image_rel_path"]

    # Start with stratified picks by subset x legs, then fill remainder randomly.
    base = df.sort_values(["subset", "legs", "style"], kind="stable")
    strat_idx = (
        base.groupby(["subset", "legs"], group_keys=False)
        .head(3)
        .index.tolist()
    )
    sample_df = df.loc[strat_idx].copy()
    if len(sample_df) < sample_n:
        remaining = sample_n - len(sample_df)
        extra_pool = df.drop(index=strat_idx, errors="ignore")
        if len(extra_pool) > 0:
            extra = extra_pool.sample(n=min(remaining, len(extra_pool)), random_state=20260322)
            sample_df = pd.concat([sample_df, extra], ignore_index=False)
    sample_df = sample_df.head(sample_n).reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "samples"
    ws.append(ws_cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for i, row in sample_df.iterrows():
        excel_row = i + 2
        img_abs = (ROOT / str(row["image_rel_path"])).resolve()
        ws.append(
            [
                "",
                row["subset"],
                row["species"],
                row["style"],
                int(row["legs"]),
                row["prompt"],
                row["answer"],
                row["image_rel_path"],
            ]
        )
        ws.row_dimensions[excel_row].height = 130
        xl_img = XLImage(str(img_abs))
        xl_img.width = 120
        xl_img.height = 120
        ws.add_image(xl_img, f"A{excel_row}")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 95
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 48

    for r in range(2, len(sample_df) + 2):
        for c in ["B", "C", "D", "E", "G"]:
            ws[f"{c}{r}"].alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws[f"F{r}"].alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        ws[f"H{r}"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    wb.save(xlsx_path)


def main() -> int:
    args = _parse_args()

    n_total = int(args.n_total)
    if n_total <= 0:
        raise ValueError("--n-total must be > 0")

    out_parquet = (ROOT / args.out_parquet).resolve()
    out_summary = (ROOT / args.out_summary).resolve()
    out_image_dir = (ROOT / args.out_image_dir).resolve()
    sample_xlsx = (ROOT / args.sample_xlsx).resolve()

    # Clean image output folder before regeneration.
    _clean_dir(out_image_dir)

    rows = _build_rows(
        n_total=n_total,
        image_size=int(args.image_size),
        aa=int(args.aa),
        out_image_dir=out_image_dir,
        seed=int(args.seed),
    )
    df = pd.DataFrame(rows)
    if not args.no_shuffle:
        df = df.sample(frac=1.0, random_state=int(args.seed)).reset_index(drop=True)

    out_parquet.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out_parquet, index=False)

    # Build embedded-image sample workbook.
    _build_sample_excel(df, xlsx_path=sample_xlsx, sample_n=int(args.sample_n))

    summary = {
        "out_parquet": str(out_parquet),
        "out_image_dir": str(out_image_dir),
        "sample_xlsx": str(sample_xlsx),
        "n_total": int(len(df)),
        "counts_by_subset": df.groupby("subset").size().to_dict(),
        "counts_by_subset_legs": (
            df.groupby(["subset", "legs"], as_index=False).size().to_dict(orient="records")
        ),
        "counts_by_style": (
            df.groupby(["subset", "style"], as_index=False).size().to_dict(orient="records")
        ),
        "seed": int(args.seed),
        "shuffle": bool(not args.no_shuffle),
        "prompts": {
            "shared": COMMON_PROMPT,
        },
    }
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    out_summary.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
